import openpyxl

wb = openpyxl.load_workbook('example.xlsx') # load your own .xlsx file
wb_new = openpyxl.Workbook()

sheet = wb.active
sheet_new = wb_new.active

for x in range(1, sheet.max_column + 1):
    for y in range(1, sheet.max_row + 1):
        sheet_new.cell(row=x, column=y).value = sheet.cell(row=y, column=x).value

wb_new.save("abc.xlsx")
print("Process Completed")
