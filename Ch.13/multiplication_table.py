import openpyxl
from openpyxl.styles import Font
import sys

wb = openpyxl.Workbook()
sheet = wb.active

fontObj = Font(bold=True)

try:
    n = int(sys.argv[1])
except ValueError:
    print("n must be number")
    sys.exit(-1)

# set values for rows & column, start with 1 to make sure A1 is empty
for i in range(1, n + 1):
    sheet['A' + str(i + 1)] = i
    sheet.cell(row=1, column=i + 1).value = i

    # setting the font to bold
    sheet['A' + str(i + 1)].font = fontObj
    sheet.cell(row=1, column=i + 1).font = fontObj

for j in range(1, n + 1):
    for k in range(1, n + 1):
        sheet.cell(row=j + 1, column=k + 1).value = j * k

file_name = 'multiplication_table_' + str(n) + "*" + str(n) + '.xlsx'

wb.save(file_name)
print('Process Completed')
