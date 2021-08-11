import openpyxl

wb = openpyxl.load_workbook('abc.xlsx') # load your own .xlsx file
sheet = wb.active

files = ["1.txt", "2.txt", "3.txt", "4.txt"]

for i in range(len(files)):
    file = open(files[i], "w")
    for j in range(sheet.max_row):
        c = i+1
        r = j+1
        file.write(str(sheet.cell(row=r, column=c).value))
        file.write(" ")
file.close()

print("Completed")




