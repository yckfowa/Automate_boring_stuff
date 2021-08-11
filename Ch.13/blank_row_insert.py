import openpyxl
import sys

try:
    n = int(sys.argv[1])  # which row to start inserting
    m = int(sys.argv[2])  # how many rows to be inserted
    file = str(sys.argv[3])  # file to be modified

except ValueError:
    print("n, m should be number, last arg should be str")
    sys.exit(-1)

wb = openpyxl.load_workbook('example.xlsx') # load your own .xlsx file
wb_new = openpyxl.Workbook()

sheet = wb.active
sheet_new = wb_new.active

for row_obj in sheet.rows:
    for cell_obj in row_obj:
        if cell_obj.row < n:
            sheet_new.cell(row=cell_obj.row, column=int(cell_obj.column)).value = cell_obj.value
        else:
            sheet_new.cell(row=cell_obj.row+m, column=int(cell_obj.column)).value = cell_obj.value

wb_new.save("example_update.xlsx")
print("Completed..")

